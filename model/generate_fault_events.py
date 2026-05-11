#!/usr/bin/env python3

from __future__ import annotations

import argparse
import math
import random
from pathlib import Path

from openpyxl import load_workbook


CODEWORD_WIDTH = 39
ADDR_WIDTH = 4
DEPTH = 1 << ADDR_WIDTH

DEFAULT_TOTAL_CYCLES = 1300


def baseline_events() -> list[tuple[int, int, int]]:
    """
    Базовый детерминированный сценарий сбойных событий.

    Формат события:
        модельный такт, адрес слова, номер бита в кодовом слове.
    """
    return [
        (120, 3, 5),
        (260, 7, 10),
        (430, 0, 2),
        (450, 0, 4),
        (740, 5, 9),
        (760, 5, 14),
        (980, 9, 1),
        (1120, 2, 20),
    ]


def read_upsets_xlsx(input_path: Path) -> list[float]:
    """
    Читает временной ряд из Excel-файла.

    Ожидаемая структура файла:
        столбец B: time
        столбец C: upsets

    Первая строка содержит заголовки.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active

    values: list[float] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue

        raw_value = row[2]

        if raw_value is None:
            continue

        value = float(raw_value)

        if not math.isfinite(value):
            continue

        if value < 0.0:
            value = 0.0

        values.append(value)

    workbook.close()

    if not values:
        raise ValueError(f"No usable upsets values found in {input_path}")

    return values


def select_window(
    values: list[float],
    start_index: int,
    window_size: int,
) -> list[float]:
    if start_index < 0:
        raise ValueError("start_index must be non-negative")

    if window_size <= 0:
        raise ValueError("window_size must be positive")

    end_index = start_index + window_size

    if end_index > len(values):
        raise ValueError(
            f"Requested window [{start_index}, {end_index}) exceeds "
            f"available series length {len(values)}"
        )

    return values[start_index:end_index]


def weighted_cycle_from_series(
    rng: random.Random,
    weights: list[float],
    total_cycles: int,
) -> int:
    """
    Выбирает такт моделирования с вероятностью,
    пропорциональной значению временного ряда.

    Индекс временного ряда сначала выбирается по весу,
    затем равномерно отображается внутрь соответствующего
    участка модельного времени.
    """
    total_weight = sum(weights)

    if total_weight <= 0.0:
        hour_index = rng.randrange(len(weights))
    else:
        threshold = rng.random() * total_weight
        cumulative = 0.0
        hour_index = 0

        for index, weight in enumerate(weights):
            cumulative += weight
            if cumulative >= threshold:
                hour_index = index
                break

    cycle_start = (hour_index * total_cycles) // len(weights)
    cycle_end = ((hour_index + 1) * total_cycles) // len(weights)

    if cycle_end <= cycle_start:
        cycle_end = cycle_start + 1

    cycle = rng.randrange(cycle_start, cycle_end)

    if cycle >= total_cycles:
        cycle = total_cycles - 1

    return cycle


def upsets_weighted_events(
    input_path: Path,
    start_index: int,
    window_size: int,
    total_cycles: int,
    event_count: int,
    seed: int,
) -> list[tuple[int, int, int]]:
    """
    Генерирует поток сбойных событий из временного ряда upsets(t).

    Это первая нагрузочная модель:
        - число событий задаётся параметром event_count;
        - момент события выбирается с вероятностью,
          пропорциональной upsets(t);
        - адрес и бит выбираются равномерно;
        - если два события попали в один такт, второе сдвигается вправо.

    Для статьи эту модель затем можно расширить до пуассоновского
    процесса с параметром, зависящим от upsets(t).
    """
    values = read_upsets_xlsx(input_path)
    window = select_window(values, start_index, window_size)

    rng = random.Random(seed)

    used_cycles: set[int] = set()
    events: list[tuple[int, int, int]] = []

    for _ in range(event_count):
        cycle = weighted_cycle_from_series(rng, window, total_cycles)

        # Если два события попали в один такт, сдвигаем новое событие
        # к ближайшему свободному такту.

        while cycle in used_cycles and cycle < total_cycles - 1:
            cycle += 1

        while cycle in used_cycles and cycle > 0:
            cycle -= 1

        if cycle in used_cycles:
            raise ValueError(
                "Could not place all events into unique cycles. "
                "Reduce event_count or increase total_cycles."
            )

        used_cycles.add(cycle)

        address = rng.randrange(DEPTH)
        bit_index = rng.randrange(CODEWORD_WIDTH)

        events.append((cycle, address, bit_index))

    events.sort(key=lambda item: item[0])
    return events


def validate_events(events: list[tuple[int, int, int]], total_cycles: int | None = None) -> None:
    previous_time = -1

    for index, (time_cycle, address, bit_index) in enumerate(events):
        if time_cycle < 0:
            raise ValueError(f"Event {index}: negative time_cycle={time_cycle}")

        if total_cycles is not None and time_cycle >= total_cycles:
            raise ValueError(
                f"Event {index}: time_cycle={time_cycle} is outside "
                f"total_cycles={total_cycles}"
            )

        if time_cycle < previous_time:
            raise ValueError(
                f"Event {index}: events must be sorted by time "
                f"({time_cycle} after {previous_time})"
            )

        if address < 0 or address >= DEPTH:
            raise ValueError(
                f"Event {index}: address={address} is outside memory depth {DEPTH}"
            )

        if bit_index < 0 or bit_index >= CODEWORD_WIDTH:
            raise ValueError(
                f"Event {index}: bit_index={bit_index} is outside "
                f"codeword width {CODEWORD_WIDTH}"
            )

        previous_time = time_cycle


def write_events(events: list[tuple[int, int, int]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="\n") as file:
        for time_cycle, address, bit_index in events:
            file.write(f"{time_cycle},{address},{bit_index}\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate fault event table for strategy comparison."
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=Path("tb/fault_events.csv"),
        help="Output CSV file without header. Default: tb/fault_events.csv",
    )

    parser.add_argument(
        "--scenario",
        choices=["baseline", "upsets"],
        default="baseline",
        help="Fault event scenario.",
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/upsets.xlsx"),
        help="Input Excel file for --scenario upsets.",
    )

    parser.add_argument(
        "--start-index",
        type=int,
        default=0,
        help="Start index in the upsets time series.",
    )

    parser.add_argument(
        "--window-size",
        type=int,
        default=1300,
        help="Number of time-series points used for upsets-based generation.",
    )

    parser.add_argument(
        "--total-cycles",
        type=int,
        default=DEFAULT_TOTAL_CYCLES,
        help="Total simulation cycles.",
    )

    parser.add_argument(
        "--event-count",
        type=int,
        default=8,
        help="Number of generated fault events for --scenario upsets.",
    )

    parser.add_argument(
        "--seed",
        type=int,
        default=12345,
        help="Random seed for reproducible event generation.",
    )

    args = parser.parse_args()

    if args.scenario == "baseline":
        events = baseline_events()
        validate_events(events, total_cycles=args.total_cycles)

    elif args.scenario == "upsets":
        events = upsets_weighted_events(
            input_path=args.input,
            start_index=args.start_index,
            window_size=args.window_size,
            total_cycles=args.total_cycles,
            event_count=args.event_count,
            seed=args.seed,
        )
        validate_events(events, total_cycles=args.total_cycles)

    else:
        raise ValueError(f"Unsupported scenario: {args.scenario}")

    write_events(events, args.output)

    print(f"Generated {len(events)} fault events: {args.output}")
    print(f"Scenario: {args.scenario}")

    if args.scenario == "upsets":
        print(f"Input: {args.input}")
        print(f"Start index: {args.start_index}")
        print(f"Window size: {args.window_size}")
        print(f"Total cycles: {args.total_cycles}")
        print(f"Seed: {args.seed}")


if __name__ == "__main__":
    main()