#!/usr/bin/env python3

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import load_workbook


DEFAULT_VALUE_COLUMN = 3

BACKGROUND_WINDOW_HOURS = 30 * 24
BACKGROUND_PERCENTILE = 0.30

R_GCR = 10.7
R_SCR = 3.0


@dataclass(frozen=True)
class SeriesStats:
    count: int
    minimum: float
    maximum: float
    mean_value: float
    std_value: float
    cv2: float
    eta_theory: float
    total_sum: float


@dataclass(frozen=True)
class FullUpsetsComponents:
    proton_raw: list[float | None]
    proton_interpolated: list[float]
    proton_background: list[float]
    proton_event: list[float]
    full_upsets: list[float]


def finite_non_negative_or_none(raw_value: object) -> float | None:
    if raw_value is None:
        return None

    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        return None

    if not math.isfinite(value):
        return None

    if value < 0.0:
        return 0.0

    return value


def read_proton_upsets_xlsx(
    input_path: Path,
    value_column: int = DEFAULT_VALUE_COLUMN,
) -> list[float | None]:
    """
    Читает протонную составляющую νp(t) из Excel-файла.

    В текущем data/upsets.xlsx ожидается:
        колонка C / индекс 3: proton upsets, то есть νp(t).

    Важно:
        это НЕ полный ряд ν(t) из статьи 3.
        Полный ряд строится ниже через Gp(t), Sp(t) и множители ТЗЧ.
    """
    if value_column <= 0:
        raise ValueError("value_column must be 1-based positive integer")

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active

    values: list[float | None] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < value_column:
            values.append(None)
            continue

        values.append(finite_non_negative_or_none(row[value_column - 1]))

    workbook.close()

    if not values:
        raise ValueError(f"No rows found in {input_path}")

    valid_count = sum(1 for value in values if value is not None)

    if valid_count == 0:
        raise ValueError(f"No usable numeric values found in {input_path}")

    return values


def interpolate_missing(values: list[float | None]) -> list[float]:
    """
    Восстанавливает пропуски в часовом ряде.

    Внутренние пропуски заполняются линейной интерполяцией.
    Ведущие и хвостовые пропуски заполняются ближайшим доступным значением.

    Это позволяет сохранить календарную длину ряда, а не выбрасывать часы.
    """
    if not values:
        raise ValueError("Cannot interpolate empty series")

    valid_indices = [index for index, value in enumerate(values) if value is not None]

    if not valid_indices:
        raise ValueError("Cannot interpolate series without valid samples")

    result = [0.0 for _ in values]

    first_valid = valid_indices[0]
    last_valid = valid_indices[-1]

    first_value = values[first_valid]
    last_value = values[last_valid]

    assert first_value is not None
    assert last_value is not None

    for index in range(0, first_valid + 1):
        result[index] = first_value

    for index in range(last_valid, len(values)):
        result[index] = last_value

    for left_index, right_index in zip(valid_indices, valid_indices[1:]):
        left_value = values[left_index]
        right_value = values[right_index]

        assert left_value is not None
        assert right_value is not None

        span = right_index - left_index

        if span <= 0:
            continue

        for offset in range(span + 1):
            alpha = offset / span
            result[left_index + offset] = (
                left_value * (1.0 - alpha) + right_value * alpha
            )

    return result


def percentile_linear(sorted_values: list[float], percentile: float) -> float:
    """
    Линейная интерполяция перцентиля по отсортированному окну.

    percentile задаётся в диапазоне [0, 1].
    """
    if not sorted_values:
        raise ValueError("Cannot compute percentile of empty list")

    if percentile < 0.0 or percentile > 1.0:
        raise ValueError("percentile must be in [0, 1]")

    if len(sorted_values) == 1:
        return sorted_values[0]

    position = (len(sorted_values) - 1) * percentile
    lower_index = int(math.floor(position))
    upper_index = int(math.ceil(position))

    if lower_index == upper_index:
        return sorted_values[lower_index]

    lower_value = sorted_values[lower_index]
    upper_value = sorted_values[upper_index]
    alpha = position - lower_index

    return lower_value * (1.0 - alpha) + upper_value * alpha


def rolling_background_percentile(
    values: list[float],
    window_hours: int = BACKGROUND_WINDOW_HOURS,
    percentile: float = BACKGROUND_PERCENTILE,
) -> list[float]:
    """
    Считает фоновую протонную составляющую Gp(t).

    Используется одностороннее скользящее окно:
        [max(0, t - window_hours + 1), t]

    Для первых точек используется доступный префикс ряда.
    """
    if window_hours <= 0:
        raise ValueError("window_hours must be positive")

    background: list[float] = []

    for index in range(len(values)):
        start = max(0, index - window_hours + 1)
        window = sorted(values[start : index + 1])
        background.append(percentile_linear(window, percentile))

    return background


def build_full_upsets_components(
    proton_raw: list[float | None],
    window_hours: int = BACKGROUND_WINDOW_HOURS,
    background_percentile: float = BACKGROUND_PERCENTILE,
    r_gcr: float = R_GCR,
    r_scr: float = R_SCR,
) -> FullUpsetsComponents:
    """
    Строит полный ряд ν(t) из протонной составляющей νp(t).

    Используется формула статьи 3:

        Gp(t) = rolling_percentile_30%(νp)
        Sp(t) = max(νp(t) - Gp(t), 0)
        ν(t)  = Gp(t) * (1 + R_GCR) + Sp(t) * (1 + R_SCR)

    где:
        R_GCR = 10.7
        R_SCR = 3.0
    """
    proton = interpolate_missing(proton_raw)
    background = rolling_background_percentile(
        values=proton,
        window_hours=window_hours,
        percentile=background_percentile,
    )

    event_component: list[float] = []
    full_upsets: list[float] = []

    for proton_value, background_value in zip(proton, background):
        event_value = proton_value - background_value

        if event_value < 0.0:
            event_value = 0.0

        full_value = (
            background_value * (1.0 + r_gcr)
            + event_value * (1.0 + r_scr)
        )

        event_component.append(event_value)
        full_upsets.append(full_value)

    return FullUpsetsComponents(
        proton_raw=proton_raw,
        proton_interpolated=proton,
        proton_background=background,
        proton_event=event_component,
        full_upsets=full_upsets,
    )


def load_full_upsets_series(
    input_path: Path,
    value_column: int = DEFAULT_VALUE_COLUMN,
) -> list[float]:
    """
    Возвращает полный ряд ν(t), а не сырую протонную составляющую.
    """
    proton_raw = read_proton_upsets_xlsx(
        input_path=input_path,
        value_column=value_column,
    )
    components = build_full_upsets_components(proton_raw)
    return components.full_upsets


def load_full_upsets_components(
    input_path: Path,
    value_column: int = DEFAULT_VALUE_COLUMN,
) -> FullUpsetsComponents:
    proton_raw = read_proton_upsets_xlsx(
        input_path=input_path,
        value_column=value_column,
    )
    return build_full_upsets_components(proton_raw)


def compute_stats(values: list[float]) -> SeriesStats:
    if not values:
        raise ValueError("Cannot compute stats of empty series")

    mean_value = mean(values)
    std_value = pstdev(values)
    cv2 = (std_value / mean_value) ** 2 if mean_value > 0.0 else 0.0

    return SeriesStats(
        count=len(values),
        minimum=min(values),
        maximum=max(values),
        mean_value=mean_value,
        std_value=std_value,
        cv2=cv2,
        eta_theory=1.0 + cv2,
        total_sum=sum(values),
    )


def print_stats(title: str, stats: SeriesStats) -> None:
    print(title)
    print(f"  count       = {stats.count}")
    print(f"  min         = {stats.minimum:.12g}")
    print(f"  max         = {stats.maximum:.12g}")
    print(f"  mean        = {stats.mean_value:.12g}")
    print(f"  std         = {stats.std_value:.12g}")
    print(f"  CV^2        = {stats.cv2:.12g}")
    print(f"  eta=1+CV^2  = {stats.eta_theory:.12g}")
    print(f"  sum         = {stats.total_sum:.12g}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Inspect proton and full upset-rate time series."
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=Path("data/upsets.xlsx"),
        help="Input Excel file with proton upset component.",
    )

    parser.add_argument(
        "--value-column",
        type=int,
        default=DEFAULT_VALUE_COLUMN,
        help="1-based column index with proton upsets. Default: 3.",
    )

    args = parser.parse_args()

    components = load_full_upsets_components(
        input_path=args.input,
        value_column=args.value_column,
    )

    proton_valid = [
        value for value in components.proton_raw
        if value is not None
    ]

    missing_count = sum(
        1 for value in components.proton_raw
        if value is None
    )

    print(f"Input: {args.input}")
    print(f"Rows: {len(components.proton_raw)}")
    print(f"Missing / non-numeric rows: {missing_count}")
    print("")

    print_stats(
        "Raw proton component νp(t), valid values only:",
        compute_stats(proton_valid),
    )
    print("")

    print_stats(
        "Interpolated proton component νp(t):",
        compute_stats(components.proton_interpolated),
    )
    print("")

    print_stats(
        "Background proton component Gp(t):",
        compute_stats(components.proton_background),
    )
    print("")

    print_stats(
        "Event proton component Sp(t):",
        compute_stats(components.proton_event),
    )
    print("")

    print_stats(
        "Full upset-rate series ν(t) with heavy charged particles:",
        compute_stats(components.full_upsets),
    )


if __name__ == "__main__":
    main()